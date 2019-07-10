import traceback
import argparse
import inspect
import Spartacus.Database
import openpyxl
import multiprocessing

import workers.custom_exceptions
import workers.compare_schemas
import workers.compare_tables
import workers.compare_tables_columns
import workers.compare_tables_pks
import workers.compare_tables_fks
import workers.compare_tables_uniques
import workers.compare_tables_checks
import workers.compare_tables_excludes
import workers.compare_indexes
import workers.compare_tables_rules
import workers.compare_tables_triggers
import workers.compare_sequences
import workers.compare_views
import workers.compare_mviews
import workers.compare_functions
import workers.compare_trigger_functions
import workers.compare_procedures
import workers.compare_tables_data


def create_sheet(p_worksheet_dict=None, p_workbook=None, p_sheet_key=None):
    """Used to create a sheet when its first row is added.

        Args:
            p_worksheet_dict (dict): dict containing all sheets that were already created. Defaults to None.
            p_workbook (openpyxl.workbook.workbook.Workbook): the workbook where sheet will be created. Defaults to None.
            p_sheet_key (str): the key of the sheet that is being created. Defaults to None.
                Notes: must be one of:
                    - schemas
                    - tables
                    - tables_columns
                    - tables_pks
                    - tables_fks
                    - tables_uniques
                    - tables_checks
                    - tables_excludes
                    - indexes
                    - tables_data
                    - tables_rules
                    - tables_triggers
                    - sequences
                    - views
                    - mviews
                    - functions
                    - trigger_functions
                    - procedures

        Raises:
            workers.custom_exceptions.InvalidParameterTypeException.
            workers.custom_exceptions.InvalidParameterValueException.
    """

    if not isinstance(p_worksheet_dict, dict):
        raise workers.custom_exceptions.InvalidParameterTypeException('"p_worksheet_dict" parameter must be a "dict" instance.', p_worksheet_dict)

    if not isinstance(p_workbook, openpyxl.workbook.workbook.Workbook):
        raise workers.custom_exceptions.InvalidParameterTypeException('"p_workbook" parameter must be an "openpyxl.workbook.workbook.Workbook" instance.', p_workbook)

    if not isinstance(p_sheet_key, str):
        raise workers.custom_exceptions.InvalidParameterTypeException('"p_sheet_key" parameter must be a "str" instance.', p_sheet_key)

    if p_sheet_key not in ['schemas', 'tables', 'tables_columns', 'tables_pks', 'tables_fks', 'tables_uniques', 'tables_checks', 'tables_excludes', 'indexes', 'tables_data', 'tables_rules', 'tables_triggers', 'sequences', 'views', 'mviews', 'functions', 'trigger_functions', 'procedures']:
        raise workers.custom_exceptions.InvalidParameterValueException(
            '"p_sheet_key" parameter must be one between: schemas, tables, tables_columns, tables_pks, tables_fks, tables_uniques, tables_checks, tables_excludes, indexes, tables_data, tables_rules, tables_triggers, sequences, views, mviews, functions, trigger_functions, procedures.',
            p_sheet_key
        )

    if p_sheet_key not in p_worksheet_dict:
        v_worksheet = p_workbook.create_sheet(p_sheet_key)

        if p_sheet_key == 'schemas':
            v_worksheet.append(['schema_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'tables':
            v_worksheet.append(['schema_name', 'table_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'tables_columns':
            v_worksheet.append(['schema_name', 'table_name', 'column_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'tables_pks':
            v_worksheet.append(['schema_name', 'table_name', 'pk_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'tables_fks':
            v_worksheet.append(['schema_name', 'table_name', 'fk_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'tables_uniques':
            v_worksheet.append(['schema_name', 'table_name', 'unique_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'tables_checks':
            v_worksheet.append(['schema_name', 'table_name', 'check_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'tables_exludes':
            v_worksheet.append(['schema_name', 'table_name', 'exclude_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'indexes':
            v_worksheet.append(['schema_name', 'index_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'tables_data':
            v_worksheet.append(['schema_name', 'table_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'tables_rules':
            v_worksheet.append(['schema_name', 'table_name', 'rule_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'tables_triggers':
            v_worksheet.append(['schema_name', 'table_name', 'trigger_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'sequences':
            v_worksheet.append(['schema_name', 'sequence_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'views':
            v_worksheet.append(['schema_name', 'view_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'mviews':
            v_worksheet.append(['schema_name', 'mview_name', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'functions':
            v_worksheet.append(['schema_name', 'function_id', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'trigger_functions':
            v_worksheet.append(['schema_name', 'trigger_function_id', 'key', 'status', 'diff', 'sql_to_fix'])
        elif p_sheet_key == 'procedures':
            v_worksheet.append(['schema_name', 'procedure_id', 'key', 'status', 'diff', 'sql_to_fix'])

        p_worksheet_dict[p_sheet_key] = v_worksheet


if __name__ == '__main__':
    try:
        v_parser = argparse.ArgumentParser(
            epilog=inspect.cleandoc(
                doc='''\
                    Script used to compare databases and find the differences between them, if any.
                    Found differences will be written to a given output xlsx file.
                '''
            )
        )

        v_parser.add_argument(
            '-o',
            '--output-file',
            dest='output_file',
            help='Full path to the output file. Must have .xlsx extension.',
            type=str,
            required=True
        )

        v_parser.add_argument(
            '-b',
            '--block-size',
            dest='block_size',
            help='Number of data records that the comparer will deal with at the same time in each opened process. Used to reduce memory usage.',
            type=int,
            required=True
        )

        v_parser.add_argument(
            '-s',
            '--source-database-connection',
            dest='source_database_connection',
            help='Connection string to the source database. Has the following structure: HOST:PORT:DATABASE:USER:PASSWORD. You can leave password empty if it is already in you .pgpass file.',
            type=str,
            required=True
        )

        v_parser.add_argument(
            '-t',
            '--target-database-connection',
            dest='target_database_connection',
            help='Connection string to the target database. Has the following structure: HOST:PORT:DATABASE:USER:PASSWORD. You can leave password empty if it is already in you .pgpass file.',
            type=str,
            required=True
        )

        v_options = v_parser.parse_args()

        #Get databases credentials
        v_source_params = v_options.source_database_connection.split(':')
        v_target_params = v_options.target_database_connection.split(':')

        #Create output file and corresponding sheets
        v_output_workbook = openpyxl.Workbook(write_only=True)
        v_output_worksheet_dict = {}

        #Open a process pool and create tasks to be run in parallel
        v_process_pool = multiprocessing.Pool(multiprocessing.cpu_count())
        v_result_list = []
        v_task_list = []

        v_task_list += workers.compare_schemas.get_compare_schemas_tasks()
        v_task_list += workers.compare_tables.get_compare_tables_tasks()
        v_task_list += workers.compare_tables_columns.get_compare_tables_columns_tasks()
        v_task_list += workers.compare_tables_pks.get_compare_tables_pks_tasks()
        v_task_list += workers.compare_tables_fks.get_compare_tables_fks_tasks()
        v_task_list += workers.compare_tables_uniques.get_compare_tables_uniques_tasks()
        v_task_list += workers.compare_tables_checks.get_compare_tables_checks_tasks()
        v_task_list += workers.compare_tables_excludes.get_compare_tables_excludes_tasks()
        v_task_list += workers.compare_indexes.get_compare_indexes_tasks()
        v_task_list += workers.compare_tables_rules.get_compare_tables_rules_tasks()
        v_task_list += workers.compare_tables_triggers.get_compare_tables_triggers_tasks()
        v_task_list += workers.compare_sequences.get_compare_sequences_tasks()
        v_task_list += workers.compare_views.get_compare_views_tasks()
        v_task_list += workers.compare_mviews.get_compare_mviews_tasks()
        v_task_list += workers.compare_functions.get_compare_functions_tasks()
        v_task_list += workers.compare_trigger_functions.get_compare_trigger_functions_tasks()
        v_task_list += workers.compare_procedures.get_compare_procedures_tasks()

        v_task_list += workers.compare_tables_data.get_compare_tables_data_tasks(
            p_database_1=Spartacus.Database.PostgreSQL(
                p_host=v_source_params[0],
                p_port=v_source_params[1],
                p_service=v_source_params[2],
                p_user=v_source_params[3],
                p_password=v_source_params[4],
                p_application_name='compare_databases'
            ),
            p_database_2=Spartacus.Database.PostgreSQL(
                p_host=v_target_params[0],
                p_port=v_target_params[1],
                p_service=v_target_params[2],
                p_user=v_target_params[3],
                p_password=v_target_params[4],
                p_application_name='compare_databases'
            ),
            p_block_size=v_options.block_size
        )

        v_manager = multiprocessing.Manager()
        v_queue = v_manager.Queue()
        v_is_sending_data_array = v_manager.Array('b', [True] * len(v_task_list))

        for i in range(len(v_task_list)):
            v_task = v_task_list[i]

            v_task['kwds']['p_database_1'] = Spartacus.Database.PostgreSQL(
                p_host=v_source_params[0],
                p_port=v_source_params[1],
                p_service=v_source_params[2],
                p_user=v_source_params[3],
                p_password=v_source_params[4],
                p_application_name='compare_databases'
            )

            v_task['kwds']['p_database_2'] = Spartacus.Database.PostgreSQL(
                p_host=v_target_params[0],
                p_port=v_target_params[1],
                p_service=v_target_params[2],
                p_user=v_target_params[3],
                p_password=v_target_params[4],
                p_application_name='compare_databases'
            )

            v_task['kwds']['p_block_size'] = v_options.block_size
            v_task['kwds']['p_queue'] = v_queue
            v_task['kwds']['p_is_sending_data_array'] = v_is_sending_data_array
            v_task['kwds']['p_worker_index'] = i

            v_result_list.append(
                v_process_pool.apply_async(
                    func=v_task['function'],
                    kwds=v_task['kwds']
                )
            )

        #While any task is still active, try to get data
        while any(v_is_sending_data_array):
            v_data = v_queue.get()

            if v_data is not None:
                if v_data['type'] not in v_output_worksheet_dict:
                    create_sheet(p_worksheet_dict=v_output_worksheet_dict, p_workbook=v_output_workbook, p_sheet_key=v_data['type'])

                v_output_worksheet_dict[v_data['type']].append(v_data['row'])

        v_process_pool.close()
        v_process_pool.join()

        #Write output file
        v_output_workbook.save(v_options.output_file)

        #If any exception in any task
        if not all([v_result.successful() for v_result in v_result_list]):
            print('Some exception has occurred in the subprocesses. Please, check the exceptions below:')

            for v_result in v_result_list:
                if not v_result.successful():
                    try:
                        v_result.get()
                    except Exception:
                        print(traceback.format_exc())
    except Exception:
        print(traceback.format_exc())
